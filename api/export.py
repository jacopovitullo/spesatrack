import io
import csv
from datetime import date
from calendar import monthrange
from flask import Blueprint, request, send_file, jsonify
from db.client import get_client

export_bp = Blueprint("export", __name__)


def _error(msg: str, code: int = 400):
    return jsonify({"error": msg}), code


def _fetch_spese(mese=None, anno=None, tutto=False):
    db = get_client()
    query = db.table("spese").select("*, categorie(nome, icona)").order("data", desc=False)

    if not tutto and mese and anno:
        giorni = monthrange(int(anno), int(mese))[1]
        query = query.gte("data", f"{anno}-{int(mese):02d}-01").lte("data", f"{anno}-{int(mese):02d}-{giorni:02d}")

    return query.execute().data or []


@export_bp.route("/export/csv", methods=["GET"])
def export_csv():
    try:
        mese = request.args.get("mese")
        anno = request.args.get("anno")
        tutto = request.args.get("tutto", "false").lower() == "true"
        spese = _fetch_spese(mese, anno, tutto)

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Data", "Descrizione", "Categoria", "Importo", "Fonte", "Note"])
        for s in spese:
            cat = s.get("categorie") or {}
            writer.writerow([
                s.get("data", ""),
                s.get("descrizione", ""),
                cat.get("nome", ""),
                s.get("importo", 0),
                s.get("fonte", ""),
                s.get("note", ""),
            ])

        output.seek(0)
        nome_file = f"spese_{mese or 'storico'}_{anno or date.today().year}.csv"
        return send_file(
            io.BytesIO(output.getvalue().encode("utf-8-sig")),
            mimetype="text/csv",
            as_attachment=True,
            download_name=nome_file,
        )
    except Exception as e:
        return _error(str(e), 500)


@export_bp.route("/export/excel", methods=["GET"])
def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        mese = request.args.get("mese")
        anno = request.args.get("anno")
        tutto = request.args.get("tutto", "false").lower() == "true"
        spese = _fetch_spese(mese, anno, tutto)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Spese"

        # Intestazione
        intestazione = ["Data", "Descrizione", "Categoria", "Importo (€)", "Fonte", "Note"]
        ws.append(intestazione)
        for col, cell in enumerate(ws[1], 1):
            cell.font = Font(bold=True, color="000000")
            cell.fill = PatternFill("solid", fgColor="C8FF00")
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = 18

        # Dati
        totale = 0.0
        for s in spese:
            cat = s.get("categorie") or {}
            importo = float(s.get("importo", 0))
            totale += importo
            ws.append([
                s.get("data", ""),
                s.get("descrizione", ""),
                cat.get("nome", ""),
                importo,
                s.get("fonte", ""),
                s.get("note", ""),
            ])

        # Riga totale
        riga_tot = ws.max_row + 1
        ws.cell(riga_tot, 1, "TOTALE").font = Font(bold=True)
        ws.cell(riga_tot, 4, totale).font = Font(bold=True)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        nome_file = f"spese_{mese or 'storico'}_{anno or date.today().year}.xlsx"
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=nome_file,
        )
    except Exception as e:
        return _error(str(e), 500)


@export_bp.route("/export/pdf", methods=["GET"])
def export_pdf():
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet

        mese = request.args.get("mese")
        anno = request.args.get("anno")
        tutto = request.args.get("tutto", "false").lower() == "true"
        spese = _fetch_spese(mese, anno, tutto)

        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()
        elementi = []

        # Titolo
        periodo = f"{mese}/{anno}" if mese and anno else "Storico completo"
        elementi.append(Paragraph(f"SpesaTrack — {periodo}", styles["Title"]))
        elementi.append(Spacer(1, 0.5*cm))

        # Tabella
        dati = [["Data", "Descrizione", "Categoria", "Importo", "Note"]]
        totale = 0.0
        for s in spese:
            cat = s.get("categorie") or {}
            importo = float(s.get("importo", 0))
            totale += importo
            dati.append([
                s.get("data", ""),
                s.get("descrizione", "")[:40],
                cat.get("nome", ""),
                f"€{importo:.2f}",
                (s.get("note") or "")[:30],
            ])
        dati.append(["", "TOTALE", "", f"€{totale:.2f}", ""])

        tabella = Table(dati, colWidths=[2.5*cm, 6*cm, 3.5*cm, 2.5*cm, 3.5*cm])
        tabella.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#c8ff00")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#f5f5f5")]),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
            ("ALIGN", (3, 0), (3, -1), "RIGHT"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
        ]))
        elementi.append(tabella)
        doc.build(elementi)

        output.seek(0)
        nome_file = f"spese_{mese or 'storico'}_{anno or date.today().year}.pdf"
        return send_file(output, mimetype="application/pdf", as_attachment=True, download_name=nome_file)
    except Exception as e:
        return _error(str(e), 500)
